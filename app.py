"""
Controle de Produtividade Banho — Pintura Eletrostática
Fluxo: Preparação (operador) -> Fila do banho -> Banho (operador) -> Concluído
(sem etapa de validação do líder)

Banco: PostgreSQL (Railway). Local sem DATABASE_URL -> SQLite.
"""
import os
import io
import csv
from datetime import datetime, timedelta

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, jsonify, send_file
)
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import create_engine, Column, Integer, String, DateTime, Float, Text
from sqlalchemy.orm import declarative_base, sessionmaker, scoped_session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'troque-esta-chave-em-producao')

DATABASE_URL = os.environ.get('DATABASE_URL', '')
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
if not DATABASE_URL:
    DATABASE_URL = 'sqlite:///dados_local.db'

engine = create_engine(DATABASE_URL, pool_pre_ping=True, pool_recycle=280)
Session = scoped_session(sessionmaker(bind=engine))
Base = declarative_base()

TOTAL_CESTOS = 19  # cestos físicos numerados de 1 a 19

PROCESSOS = [
    "AÇO SEM OXIDAÇÃO",
    "AÇO COM OXIDAÇÃO",
    "ALUMÍNIO",
    "MINIMIZADO SEM OXIDAÇÃO",
    "MINIMIZADO COM OXIDAÇÃO",
    "INOX",
]

# Estados do card
ST_PREPARANDO = 'PREPARANDO'      # operador iniciou, enchendo o cesto (cronômetro rodando)
ST_PREENCHER = 'PREENCHER'        # tempo parado, aguardando preenchimento dos dados
ST_FILA_BANHO = 'FILA_BANHO'      # preparação concluída, aguardando banho
ST_EM_BANHO = 'EM_BANHO'          # operador de banho iniciou
ST_CONCLUIDO = 'CONCLUIDO'        # finalizado
ESTADOS_ATIVOS = (ST_PREPARANDO, ST_PREENCHER, ST_FILA_BANHO, ST_EM_BANHO)


# ─────────────────────────────────────────────────────────────────────────────
# Modelos
# ─────────────────────────────────────────────────────────────────────────────
class Usuario(Base):
    __tablename__ = 'usuarios'
    id = Column(Integer, primary_key=True)
    login = Column(String(50), unique=True, nullable=False)
    nome = Column(String(120), nullable=False)
    senha_hash = Column(String(255), nullable=False)
    perfil = Column(String(20), nullable=False)  # admin, prep, banho

    def to_dict(self):
        return {'id': self.id, 'login': self.login, 'nome': self.nome, 'perfil': self.perfil}


class ItemMestre(Base):
    """Lista mestra SAP: Ordem(OP) -> Material(código), Texto breve, Quantidade total."""
    __tablename__ = 'itens_mestre'
    id = Column(Integer, primary_key=True)
    ordem = Column(String(60), unique=True, nullable=False, index=True)  # OP
    material = Column(String(60), default='')                           # código
    texto_breve = Column(String(255), default='')                       # descrição
    quantidade = Column(Integer, default=0)                             # quantidade total

    def to_dict(self):
        return {'ordem': self.ordem, 'material': self.material,
                'texto_breve': self.texto_breve, 'quantidade': self.quantidade}


class Card(Base):
    __tablename__ = 'cards'
    id = Column(Integer, primary_key=True)
    estado = Column(String(20), nullable=False, index=True)

    numero_cesto = Column(Integer, nullable=False)
    processo = Column(String(60), default='')
    tipo = Column(String(20), default='Normal')   # Normal / Retrabalho

    # dados SAP
    ordem = Column(String(60), default='')        # OP
    material = Column(String(60), default='')     # código
    texto_breve = Column(String(255), default='')
    quantidade = Column(Integer, default=0)
    observacao = Column(Text, default='')

    operador_prep = Column(String(120), default='')
    operador_banho = Column(String(120), default='')

    prep_inicio = Column(DateTime)
    prep_fim = Column(DateTime)
    prep_minutos = Column(Float, default=0)

    # pausa (café, ginástica laboral...) — só na preparação
    pausado = Column(Integer, default=0)        # 0/1
    pausa_inicio = Column(DateTime)             # quando a pausa atual começou
    pausa_acumulada_seg = Column(Integer, default=0)  # total já pausado (segundos)

    banho_inicio = Column(DateTime)
    banho_fim = Column(DateTime)
    banho_minutos = Column(Float, default=0)

    criado_em = Column(DateTime, default=datetime.utcnow)

    def to_dict(self):
        def fmt(dt):
            if not dt:
                return ''
            return (dt - timedelta(hours=3)).strftime('%d/%m/%Y %H:%M:%S')

        def iso(dt):
            return dt.isoformat() + 'Z' if dt else ''
        return {
            'id': self.id, 'estado': self.estado,
            'numero_cesto': self.numero_cesto,
            'processo': self.processo, 'tipo': self.tipo,
            'ordem': self.ordem, 'material': self.material,
            'texto_breve': self.texto_breve, 'quantidade': self.quantidade,
            'observacao': self.observacao or '',
            'operador_prep': self.operador_prep, 'operador_banho': self.operador_banho,
            'prep_inicio': fmt(self.prep_inicio), 'prep_fim': fmt(self.prep_fim),
            'prep_minutos': round(self.prep_minutos or 0, 1),
            'banho_inicio': fmt(self.banho_inicio), 'banho_fim': fmt(self.banho_fim),
            'banho_minutos': round(self.banho_minutos or 0, 1),
            'prep_inicio_iso': iso(self.prep_inicio),
            'banho_inicio_iso': iso(self.banho_inicio),
            'pausado': bool(self.pausado),
            'pausa_inicio_iso': iso(self.pausa_inicio),
            'pausa_acumulada_seg': self.pausa_acumulada_seg or 0,
            # data simples (yyyy-mm-dd) p/ filtros do dashboard público
            'data_ref': (self.banho_fim - timedelta(hours=3)).strftime('%Y-%m-%d') if self.banho_fim else '',
        }


# ─────────────────────────────────────────────────────────────────────────────
# Init + seed
# ─────────────────────────────────────────────────────────────────────────────
def _migrar_colunas():
    """Adiciona colunas novas em tabelas que já existem (Postgres/SQLite)."""
    from sqlalchemy import inspect, text
    insp = inspect(engine)
    if 'cards' not in insp.get_table_names():
        return
    existentes = {c['name'] for c in insp.get_columns('cards')}
    novas = {
        'pausado': 'INTEGER DEFAULT 0',
        'pausa_inicio': 'TIMESTAMP NULL',
        'pausa_acumulada_seg': 'INTEGER DEFAULT 0',
    }
    with engine.begin() as conn:
        for col, tipo in novas.items():
            if col not in existentes:
                try:
                    conn.execute(text(f'ALTER TABLE cards ADD COLUMN {col} {tipo}'))
                except Exception:
                    pass


def init_db():
    Base.metadata.create_all(engine)
    _migrar_colunas()
    db = Session()
    try:
        if db.query(Usuario).count() == 0:
            seed = [
                ('admin', 'Administrador', 'admin123', 'admin'),
                ('banho', 'Operador de Banho', 'banho123', 'banho'),
            ]
            for i in range(1, 7):
                seed.append((f'op{i}', f'Operador {i}', 'op1234', 'prep'))
            for login, nome, senha, perfil in seed:
                db.add(Usuario(login=login, nome=nome,
                               senha_hash=generate_password_hash(senha), perfil=perfil))
            db.commit()
    finally:
        db.close()


def login_required(*perfis):
    from functools import wraps

    def deco(f):
        @wraps(f)
        def wrapper(*a, **kw):
            if 'usuario' not in session:
                return redirect(url_for('login'))
            if perfis and session.get('perfil') not in perfis and session.get('perfil') != 'admin':
                return redirect(url_for('login'))
            return f(*a, **kw)
        return wrapper
    return deco


# ─────────────────────────────────────────────────────────────────────────────
# Páginas
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    erro = None
    if request.method == 'POST':
        login_u = request.form.get('usuario', '').strip()
        senha = request.form.get('senha', '')
        db = Session()
        try:
            u = db.query(Usuario).filter_by(login=login_u).first()
            if u and check_password_hash(u.senha_hash, senha):
                session['usuario'] = u.login
                session['nome'] = u.nome
                session['perfil'] = u.perfil
                destino = {'admin': 'dashboard', 'banho': 'tela_banho',
                           'prep': 'tela_prep'}.get(u.perfil, 'login')
                return redirect(url_for(destino))
            erro = 'Usuário ou senha incorretos.'
        finally:
            db.close()
    return render_template('login.html', erro=erro)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/preparacao')
@login_required('prep', 'banho')
def tela_prep():
    return render_template('prep.html', nome=session.get('nome'),
                           perfil=session.get('perfil'), processos=PROCESSOS,
                           total_cestos=TOTAL_CESTOS)


@app.route('/banho')
@login_required('banho')
def tela_banho():
    return render_template('banho.html', nome=session.get('nome'),
                           perfil=session.get('perfil'))


@app.route('/dashboard')
@login_required('admin')
def dashboard():
    return render_template('dashboard.html', nome=session.get('nome'), processos=PROCESSOS)


# Dashboard PÚBLICO (sem login) — só leitura, para a gerência
@app.route('/painel')
def painel_publico():
    return render_template('painel.html', processos=PROCESSOS)


@app.route('/admin/usuarios', methods=['GET', 'POST'])
@login_required('admin')
def admin_usuarios():
    db = Session()
    msg = None
    try:
        if request.method == 'POST':
            acao = request.form.get('acao')
            if acao == 'adicionar':
                nu = request.form.get('novo_usuario', '').strip()
                nn = request.form.get('novo_nome', '').strip()
                ns = request.form.get('nova_senha', '')
                npf = request.form.get('novo_perfil', 'prep')
                if nu and nn and ns and not db.query(Usuario).filter_by(login=nu).first():
                    db.add(Usuario(login=nu, nome=nn,
                                   senha_hash=generate_password_hash(ns), perfil=npf))
                    db.commit()
                    msg = f'Usuário {nn} adicionado.'
                elif db.query(Usuario).filter_by(login=nu).first():
                    msg = 'Esse login já existe.'
            elif acao == 'remover':
                u = db.query(Usuario).filter_by(login=request.form.get('usuario_remover')).first()
                if u and u.login != 'admin':
                    db.delete(u)
                    db.commit()
                    msg = 'Usuário removido.'
            elif acao == 'senha':
                u = db.query(Usuario).filter_by(login=request.form.get('usuario_senha')).first()
                nova = request.form.get('senha_nova', '')
                if u and nova:
                    u.senha_hash = generate_password_hash(nova)
                    db.commit()
                    msg = f'Senha de {u.nome} atualizada.'
        usuarios = [u.to_dict() for u in db.query(Usuario).order_by(Usuario.id).all()]
        return render_template('usuarios.html', usuarios=usuarios,
                               nome=session.get('nome'), msg=msg)
    finally:
        db.close()


@app.route('/admin/mestre', methods=['GET', 'POST'])
@login_required('admin')
def admin_mestre():
    db = Session()
    msg = None
    try:
        if request.method == 'POST':
            f = request.files.get('arquivo')
            if f and f.filename:
                nome = f.filename.lower()
                try:
                    linhas = []
                    if nome.endswith('.csv') or nome.endswith('.txt'):
                        raw = f.stream.read().decode('utf-8-sig', errors='replace')
                        sep = '\t' if raw.count('\t') > raw.count(';') and raw.count('\t') > raw.count(',') \
                            else (';' if raw.count(';') > raw.count(',') else ',')
                        linhas = list(csv.reader(io.StringIO(raw), delimiter=sep))
                    else:
                        wb = load_workbook(f, read_only=True, data_only=True)
                        ws = wb.active
                        for row in ws.iter_rows(values_only=True):
                            linhas.append(list(row))
                    novos, atual = importar_mestre(db, linhas)
                    msg = f'Importado: {novos} novas ordens, {atual} atualizadas.'
                except Exception as e:
                    msg = f'Erro ao importar: {e}'
        total = db.query(ItemMestre).count()
        amostra = [i.to_dict() for i in db.query(ItemMestre).limit(25).all()]
        return render_template('mestre.html', nome=session.get('nome'),
                               msg=msg, total_itens=total, amostra=amostra)
    finally:
        db.close()


def importar_mestre(db, linhas):
    """
    Formato SAP. Colunas:
    Ordem | Nº do item | Material | Texto breve material | Quantidade total | ...
    Usamos: Ordem(0), Material(2), Texto breve(3), Quantidade total(4).
    """
    novos = atual = 0
    for i, row in enumerate(linhas):
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue
        c0 = str(row[0]).strip().lower()
        if c0 in ('ordem', 'order') or 'ordem' in c0:  # cabeçalho
            continue
        ordem = str(row[0]).strip()
        if not ordem or not ordem.replace('.', '').isdigit():
            continue
        material = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
        texto = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
        try:
            qtd = int(float(row[4])) if len(row) > 4 and row[4] not in (None, '') else 0
        except (ValueError, TypeError):
            qtd = 0
        ex = db.query(ItemMestre).filter_by(ordem=ordem).first()
        if ex:
            ex.material, ex.texto_breve, ex.quantidade = material, texto, qtd
            atual += 1
        else:
            db.add(ItemMestre(ordem=ordem, material=material, texto_breve=texto, quantidade=qtd))
            novos += 1
    db.commit()
    return novos, atual


# ─────────────────────────────────────────────────────────────────────────────
# APIs — grade de cestos e fluxo
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/cestos')
@login_required('prep', 'banho')
def api_cestos():
    """Estado de cada cesto 1..19: livre ou ocupado (com dados do card ativo)."""
    db = Session()
    try:
        ativos = db.query(Card).filter(Card.estado.in_(ESTADOS_ATIVOS)).all()
        mapa = {c.numero_cesto: c for c in ativos}
        grade = []
        for n in range(1, TOTAL_CESTOS + 1):
            c = mapa.get(n)
            grade.append({'numero': n, 'ocupado': c is not None,
                          'card': c.to_dict() if c else None})
        return jsonify(grade)
    finally:
        db.close()


@app.route('/api/buscar_ordem/<path:ordem>')
@login_required('prep', 'banho')
def api_buscar_ordem(ordem):
    db = Session()
    try:
        item = db.query(ItemMestre).filter_by(ordem=ordem.strip()).first()
        if item:
            return jsonify({'encontrado': True, **item.to_dict()})
        return jsonify({'encontrado': False, 'ordem': ordem})
    finally:
        db.close()


@app.route('/api/prep/iniciar', methods=['POST'])
@login_required('prep', 'banho')
def api_prep_iniciar():
    d = request.json or {}
    try:
        numero = int(d.get('numero_cesto'))
    except (ValueError, TypeError):
        return jsonify({'sucesso': False, 'erro': 'Cesto inválido.'}), 400
    if not (1 <= numero <= TOTAL_CESTOS):
        return jsonify({'sucesso': False, 'erro': 'Cesto fora do intervalo.'}), 400
    db = Session()
    try:
        if db.query(Card).filter(Card.numero_cesto == numero,
                                 Card.estado.in_(ESTADOS_ATIVOS)).first():
            return jsonify({'sucesso': False, 'erro': f'Cesto {numero} já está em uso.'}), 400
        card = Card(estado=ST_PREPARANDO, numero_cesto=numero,
                    operador_prep=session.get('nome', ''), prep_inicio=datetime.utcnow())
        db.add(card)
        db.commit()
        return jsonify({'sucesso': True, 'id': card.id})
    finally:
        db.close()


@app.route('/api/prep/pausar', methods=['POST'])
@login_required('prep', 'banho')
def api_prep_pausar():
    """Pausa ou retoma o cronômetro de um cesto em preparação (café, ginástica...)."""
    d = request.json or {}
    db = Session()
    try:
        card = db.query(Card).get(int(d.get('id', 0)))
        if not card or card.estado != ST_PREPARANDO:
            return jsonify({'sucesso': False, 'erro': 'Cesto não está em preparação.'}), 404
        agora = datetime.utcnow()
        if card.pausado:
            # retomar: soma o tempo que ficou pausado
            if card.pausa_inicio:
                card.pausa_acumulada_seg = (card.pausa_acumulada_seg or 0) + \
                    int((agora - card.pausa_inicio).total_seconds())
            card.pausado = 0
            card.pausa_inicio = None
        else:
            # pausar
            card.pausado = 1
            card.pausa_inicio = agora
        db.commit()
        return jsonify({'sucesso': True, 'pausado': bool(card.pausado)})
    finally:
        db.close()


@app.route('/api/prep/parar', methods=['POST'])
@login_required('prep', 'banho')
def api_prep_parar():
    """Para o cronômetro (registra fim da prep) e vai para o estado de preenchimento."""
    d = request.json or {}
    db = Session()
    try:
        card = db.query(Card).get(int(d.get('id', 0)))
        if not card or card.estado != ST_PREPARANDO:
            return jsonify({'sucesso': False, 'erro': 'Cesto não está em preparação.'}), 404
        agora = datetime.utcnow()
        # se estava pausado, fecha a pausa antes de parar
        if card.pausado and card.pausa_inicio:
            card.pausa_acumulada_seg = (card.pausa_acumulada_seg or 0) + \
                int((agora - card.pausa_inicio).total_seconds())
            card.pausado = 0
            card.pausa_inicio = None
        card.prep_fim = agora
        bruto = (card.prep_fim - card.prep_inicio).total_seconds()
        liquido = max(0, bruto - (card.pausa_acumulada_seg or 0))  # desconta pausas
        card.prep_minutos = round(liquido / 60, 1)
        card.estado = ST_PREENCHER
        db.commit()
        return jsonify({'sucesso': True, 'prep_minutos': card.prep_minutos})
    finally:
        db.close()


@app.route('/api/prep/finalizar', methods=['POST'])
@login_required('prep', 'banho')
def api_prep_finalizar():
    """Salva os dados preenchidos e envia o cesto para a fila do banho."""
    d = request.json or {}
    db = Session()
    try:
        card = db.query(Card).get(int(d.get('id', 0)))
        if not card or card.estado not in (ST_PREENCHER, ST_PREPARANDO):
            return jsonify({'sucesso': False, 'erro': 'Card não encontrado.'}), 404
        # se ainda estava em PREPARANDO (parou e preencheu de uma vez), calcula o tempo
        if card.estado == ST_PREPARANDO:
            agora = datetime.utcnow()
            if card.pausado and card.pausa_inicio:
                card.pausa_acumulada_seg = (card.pausa_acumulada_seg or 0) + \
                    int((agora - card.pausa_inicio).total_seconds())
                card.pausado = 0
                card.pausa_inicio = None
            card.prep_fim = agora
            bruto = (card.prep_fim - card.prep_inicio).total_seconds()
            card.prep_minutos = round(max(0, bruto - (card.pausa_acumulada_seg or 0)) / 60, 1)
        _aplicar_campos(card, d)
        card.estado = ST_FILA_BANHO
        db.commit()
        return jsonify({'sucesso': True})
    finally:
        db.close()


@app.route('/api/card/editar', methods=['POST'])
@login_required('prep', 'banho')
def api_card_editar():
    """Edita os dados de um card já criado (qualquer estado), corrigindo erros."""
    d = request.json or {}
    db = Session()
    try:
        card = db.query(Card).get(int(d.get('id', 0)))
        if not card:
            return jsonify({'sucesso': False, 'erro': 'Card não encontrado.'}), 404
        _aplicar_campos(card, d)
        # permite corrigir tempos manualmente se enviados
        for campo in ('prep_minutos', 'banho_minutos'):
            if d.get(campo) not in (None, ''):
                try:
                    setattr(card, campo, round(float(d.get(campo)), 1))
                except (ValueError, TypeError):
                    pass
        db.commit()
        return jsonify({'sucesso': True})
    finally:
        db.close()


def _aplicar_campos(card, d):
    for campo in ('processo', 'tipo', 'ordem', 'material', 'texto_breve', 'observacao'):
        if campo in d:
            setattr(card, campo, (d.get(campo) or '').strip() if isinstance(d.get(campo), str) else d.get(campo))
    if 'quantidade' in d:
        try:
            card.quantidade = int(d.get('quantidade') or 0)
        except (ValueError, TypeError):
            card.quantidade = 0


@app.route('/api/banho/fila')
@login_required('banho')
def api_banho_fila():
    db = Session()
    try:
        fila = db.query(Card).filter_by(estado=ST_FILA_BANHO).order_by(Card.prep_fim).all()
        emb = db.query(Card).filter_by(estado=ST_EM_BANHO).order_by(Card.banho_inicio).all()
        return jsonify({'fila': [c.to_dict() for c in fila],
                        'em_banho': [c.to_dict() for c in emb]})
    finally:
        db.close()


@app.route('/api/banho/iniciar', methods=['POST'])
@login_required('banho')
def api_banho_iniciar():
    d = request.json or {}
    db = Session()
    try:
        card = db.query(Card).get(int(d.get('id', 0)))
        if not card or card.estado != ST_FILA_BANHO:
            return jsonify({'sucesso': False, 'erro': 'Card não está na fila.'}), 404
        card.banho_inicio = datetime.utcnow()
        card.operador_banho = session.get('nome', '')
        card.estado = ST_EM_BANHO
        db.commit()
        return jsonify({'sucesso': True})
    finally:
        db.close()


@app.route('/api/banho/finalizar', methods=['POST'])
@login_required('banho')
def api_banho_finalizar():
    d = request.json or {}
    db = Session()
    try:
        card = db.query(Card).get(int(d.get('id', 0)))
        if not card or card.estado != ST_EM_BANHO:
            return jsonify({'sucesso': False, 'erro': 'Card não está em banho.'}), 404
        card.banho_fim = datetime.utcnow()
        card.banho_minutos = round((card.banho_fim - card.banho_inicio).total_seconds() / 60, 1)
        card.estado = ST_CONCLUIDO   # libera o cesto na grade
        db.commit()
        return jsonify({'sucesso': True, 'banho_minutos': card.banho_minutos})
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────────────────
# Dados para dashboards (admin e público compartilham)
# ─────────────────────────────────────────────────────────────────────────────
def _coletar_dados(de=None, ate=None):
    db = Session()
    try:
        q = db.query(Card).filter_by(estado=ST_CONCLUIDO)
        cards = q.all()

        def dentro(c):
            if not c.banho_fim:
                return False
            dia = (c.banho_fim - timedelta(hours=3)).date()
            if de and dia < de:
                return False
            if ate and dia > ate:
                return False
            return True
        cards = [c for c in cards if dentro(c)]
        ativos = db.query(Card).filter(Card.estado.in_(ESTADOS_ATIVOS)).order_by(Card.id.desc()).all()

        total = len(cards)
        normais = sum(1 for c in cards if c.tipo == 'Normal')
        retrab = sum(1 for c in cards if c.tipo == 'Retrabalho')
        tp = [c.prep_minutos for c in cards if c.prep_minutos]
        tb = [c.banho_minutos for c in cards if c.banho_minutos]
        por_proc = {}
        por_dia = {}
        for c in cards:
            p = c.processo or 'Sem processo'
            por_proc[p] = por_proc.get(p, 0) + 1
            dia = (c.banho_fim - timedelta(hours=3)).strftime('%d/%m')
            por_dia[dia] = por_dia.get(dia, 0) + 1
        return {
            'total': total, 'normais': normais, 'retrabalhos': retrab,
            'em_andamento': len(ativos),
            'media_prep': round(sum(tp) / len(tp), 1) if tp else 0,
            'media_banho': round(sum(tb) / len(tb), 1) if tb else 0,
            'por_processo': por_proc, 'por_dia': por_dia,
            'ativos': [c.to_dict() for c in ativos],
            'registros': [c.to_dict() for c in sorted(cards, key=lambda x: x.id, reverse=True)[:200]],
        }
    finally:
        db.close()


def _parse_datas():
    def pd(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return None
    return pd(request.args.get('de')), pd(request.args.get('ate'))


@app.route('/api/dashboard/dados')
@login_required('admin')
def api_dashboard_dados():
    de, ate = _parse_datas()
    return jsonify(_coletar_dados(de, ate))


@app.route('/api/painel/dados')
def api_painel_dados():
    """Endpoint público — só leitura, sem dados sensíveis de login."""
    de, ate = _parse_datas()
    d = _coletar_dados(de, ate)
    return jsonify(d)


# ─────────────────────────────────────────────────────────────────────────────
# Exports Excel (pré-banho e banho separados)
# ─────────────────────────────────────────────────────────────────────────────
def _estilo_cabecalho(ws, headers):
    fill = PatternFill("solid", fgColor="0F3D5C")
    font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style='thin', color='D0D7DE')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.row_dimensions[1].height = 28


def _gerar_excel(tipo):
    db = Session()
    try:
        cards = db.query(Card).filter_by(estado=ST_CONCLUIDO).order_by(Card.id).all()
        wb = Workbook()
        ws = wb.active
        if tipo == 'prebanho':
            ws.title = 'Pre-Banho'
            headers = ['ID', 'Cesto', 'OP (Ordem)', 'Código (Material)', 'Texto breve',
                       'Qtd total', 'Processo', 'Tipo', 'Operador',
                       'Início', 'Fim', 'Tempo (min)', 'Observação']
            larg = [6, 8, 14, 16, 32, 10, 22, 12, 18, 20, 20, 12, 30]
            _estilo_cabecalho(ws, headers)
            for c in cards:
                dd = c.to_dict()
                ws.append([dd['id'], dd['numero_cesto'], dd['ordem'], dd['material'],
                           dd['texto_breve'], dd['quantidade'], dd['processo'], dd['tipo'],
                           dd['operador_prep'], dd['prep_inicio'], dd['prep_fim'],
                           dd['prep_minutos'], dd['observacao']])
        else:
            ws.title = 'Banho'
            headers = ['ID', 'Cesto', 'OP (Ordem)', 'Código (Material)', 'Texto breve',
                       'Qtd total', 'Processo', 'Tipo', 'Operador banho',
                       'Início banho', 'Fim banho', 'Tempo (min)']
            larg = [6, 8, 14, 16, 32, 10, 22, 12, 18, 20, 20, 12]
            _estilo_cabecalho(ws, headers)
            for c in cards:
                dd = c.to_dict()
                ws.append([dd['id'], dd['numero_cesto'], dd['ordem'], dd['material'],
                           dd['texto_breve'], dd['quantidade'], dd['processo'], dd['tipo'],
                           dd['operador_banho'], dd['banho_inicio'], dd['banho_fim'],
                           dd['banho_minutos']])
        for i, w in enumerate(larg, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
        ws.freeze_panes = 'A2'
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    finally:
        db.close()


@app.route('/api/download/prebanho')
@login_required('admin')
def download_prebanho():
    buf = _gerar_excel('prebanho')
    stamp = datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(buf, as_attachment=True, download_name=f'prebanho_{stamp}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/download/banho')
@login_required('admin')
def download_banho():
    buf = _gerar_excel('banho')
    stamp = datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(buf, as_attachment=True, download_name=f'banho_{stamp}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.teardown_appcontext
def remove_session(exc=None):
    Session.remove()


init_db()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
